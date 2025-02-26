from django.shortcuts import render,redirect,get_object_or_404
from .models import Product,ShoppingList
from django.http import HttpResponse
from django.conf import settings
import uuid
import openpyxl
from openpyxl.styles import Font
from django.db.models import Q
import requests
import json
from django.http import JsonResponse
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import simpleSplit
from django.utils.timezone import now

def generate_shopping_list_pdf(session_key):
    """Generates a PDF file of the current user's shopping list."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=shopping_list.pdf'

    # Create PDF Canvas
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    pdf.setTitle("Shopping List")
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(200, height - 50, "Shopping List")

    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, height - 70, f"Generated on: {now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Table Headers
    y_position = height - 100
    pdf.setFont("Helvetica-Bold", 12)
    headers = ["Product", "Quantity", "Price (KSH)", "Total Price (KSH)"]
    x_positions = [50, 200, 300, 400]

    for i, header in enumerate(headers):
        pdf.drawString(x_positions[i], y_position, header)

    # Fetch shopping items for the specific user
    shopping_items = ShoppingList.objects.filter(session_key=session_key)
    total_price = 0
    y_position -= 20
    pdf.setFont("Helvetica", 10)

    for item in shopping_items:
        if y_position < 50:  # Check if page needs a break
            pdf.showPage()
            y_position = height - 50
            pdf.setFont("Helvetica", 10)

        total_item_price = item.quantity * item.product.Price
        total_price += total_item_price

        # Ensure text doesn't overflow
        product_name = "\n".join(simpleSplit(item.product.name, "Helvetica", 10, 140))
        pdf.drawString(x_positions[0], y_position, product_name)
        pdf.drawString(x_positions[1], y_position, str(item.quantity))
        pdf.drawString(x_positions[2], y_position, f"{item.product.Price:.2f}")
        pdf.drawString(x_positions[3], y_position, f"{total_item_price:.2f}")

        y_position -= 20

    # Draw total price
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(300, y_position - 10, f"Overall Total (KSH): {total_price:.2f}")

    pdf.save()

    # Clear shopping list after generating the PDF
    shopping_items.delete()

    return response


def download_shopping_list_as_pdf(request):
    """View to generate and return the shopping list as a PDF file for the current user."""
    if not request.session.session_key:
        return HttpResponse("No active shopping session found.", status=400)

    session_key = request.session.session_key
    return generate_shopping_list_pdf(session_key)

def generate_shopping_list_xlsx(session_key):
    """Generates an XLSX file of the current user's shopping list."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Shopping List"
    
    # Define headers
    headers = ["Product", "Quantity", "Price (KSH)", "Total Price (KSH)"]
    sheet.append(headers)
    
    # Fetch shopping items for the specific user
    shopping_items = ShoppingList.objects.filter(session_key=session_key)
    total_price = 0
    
    for item in shopping_items:
        total_item_price = item.quantity * item.product.Price
        total_price += total_item_price
        sheet.append([
            item.product.name,
            item.quantity,
            item.product.Price,
            total_item_price
        ])
    
    # Add overall total row
    sheet.append(["", "", "Overall Total (KSH):", total_price])
    
    # Auto-adjust column width
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column].width = max_length + 2
    
    return workbook

def download_shopping_list_after_payment(request):
    """View to generate and return the shopping list as an XLSX file for the current user."""
    if not request.session.session_key:
        return HttpResponse("No active shopping session found.", status=400)

    session_key = request.session.session_key
    workbook = generate_shopping_list_xlsx(session_key)

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=shopping_list.xlsx'
    
    workbook.save(response)

    # Clear shopping list after download
    ShoppingList.objects.filter(session_key=session_key).delete()

    return response

def home(request):
    products = Product.objects.filter(quantity__gt=0)
    return render(request, 'home/homes.html', {
        'products':products,
    })

def shopping_list_view(request):
    # Get or create a session key for each user
    if not request.session.session_key:
        request.session.create()  # Generates a unique session key
    session_key = request.session.session_key

    # Get only the shopping items for this user
    shopping_items = ShoppingList.objects.filter(session_key=session_key)

    # Assign a unique_id if missing
    for item in shopping_items:
        if not item.unique_id:
            item.unique_id = uuid.uuid4()
            item.save()

    # Calculate total price
    total_price = sum(item.quantity * item.product.Price for item in shopping_items)

    # Add total cost per item
    for item in shopping_items:
        item.total_cost = item.quantity * item.product.Price

    context = {
        'shopping_items': shopping_items,
        'total_price': total_price,
    }

    return render(request, 'home/shopping_list.html', context)


def add_to_shopping_list(request, product_id):
    if not request.session.session_key:
        request.session.create()
    session_key = request.session.session_key

    product = Product.objects.get(id=product_id)
    shopping_item, created = ShoppingList.objects.get_or_create(
        product=product,
        session_key=session_key,  # Ensure item belongs to current user
        defaults={'quantity': 1}
    )
    
    if not created:
        shopping_item.quantity += 1  # Increment quantity if item exists
        shopping_item.save()

    return redirect('home:shopping_list')



def update_shopping_list(request, item_id):
    if request.method == "POST":
        shopping_item = get_object_or_404(ShoppingList, id=item_id)
        quantity = int(request.POST.get("quantity", 1))  
        shopping_item.quantity = quantity
        shopping_item.save()
        return redirect('home:shopping_list')

def delete_from_shopping_list(request, item_id):
    if request.method == "POST":
        shopping_item = get_object_or_404(ShoppingList, id=item_id)
        shopping_item.delete()
        return redirect('home:shopping_list')

def shopping_list_detail(request, unique_id):
    shopping_list = get_object_or_404(ShoppingList, unique_id=unique_id)
    items = shopping_list.items.all()

    context = {
        'shopping_list': shopping_list,
        'items': items,
        'total_price': shopping_list.total_price(),
    }
    return render(request, 'home/shopping_list_detail.html', context)

def detail(request,pk):
    item = get_object_or_404(Product, pk=pk)
    return render(request, 'home/detail.html', {
        'item':item,
    })


def search(request):
    query = request.GET.get('query','')
    items = []
    if query:
        query_words = query.split()
        items_query = Q()
        for word in query_words:
            items_query &= Q(name__icontains=word) | Q(description__icontains=word)
        items = Product.objects.filter(items_query)
    return render(request,'home/search.html',{
        'query':query,
        'items':items,
    }) 
    
    
def help(request):
    return render(request, 'home/help.html') 
